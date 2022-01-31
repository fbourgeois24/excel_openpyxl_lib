from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

class excel_file():

	def __init__(self, filepath, next_line=1):
		"""
			filepath est le chemin complet du fichier
			next_line est la ligne suivante à laquelle écrire, par défaut, la première
		"""
		self.filepath = filepath
		self.next_line = next_line


	def create_workbook(self):
		wb = Workbook()
		wb.save(self.filepath)

	def open(self, sheet='first'):
		self.wb = load_workbook(self.filepath)
		if sheet == 'first':
			self.ws = self.wb.active
		else:
			raise TypeError("Valeur de 'sheet' incorrecte")

	def write(self, data, from_cell=1, from_line=None):
		""" Ecriture dans le fichier, 
			Données à écrire:
				- Si données n'est pas une liste ou un tuple, on écrit une seule case
				- Si données est une liste ou un tuple à un seul niveau, on écrit une ligne
				- Si données est une liste ou un tuple à deux niveaux, on écrit plusieurs lignes
			Ligne à écrire (une ligne libre est une ligne dont la première cellule est vide):
				- si from_line est null on écrit à la ligne suivante enregistrée dans self.next_line
				- si line est à first_free, on cherche la première ligne libre
				- si line est à next_free, on cherche la ligne libre suivante après self.next_line
		"""
		
		# On défini la ligne à écrire
		if from_line is None:	
			from_line = self.next_line
		elif from_line == "first_free":
			while self.ws.cell(1, 1).value is not None:
				from_line += 1
		elif from_line == 'next_free':
			while self.ws.cell(self.next_line, 1).value is not None:
				from_line += 1
		elif from_line.isdigit():
			pass
		else:
			raise TypeError("Valeur de ligne à écrire incorrecte")

		# On défini la taille des données à écrire
		# data_size est une liste qui contiendra les deux dimentions des données à écrite [ligne, colonne]
		if type(data) not in (tuple, list):
			# Cellule unique
			data_size = [1,1]
		elif type(data[0]) not in (tuple, list):
			# On défini si c'est une liste à deux dimentions en regardant le premier élément de la liste
			# Liste à une dimention
			data_size = [1, len(data)]
		else:
			# Liste à deux dimentions
			data_size = [len(data[0]), len(data)]


		# On écrit les données en fonction des tailles définies 
		raise SyntaxError("Programmation à compléter")


	def append(self, data):
		""" Ajout d'une ou plusieurs lignes directement à la suite
			Data doit obligatoirement être un tuple ou une liste même s'il n'y a qu'un seul élément
		"""
		if type(data) not in (list, tuple):
			raise TypeError("Les données doivent être une liste ou un tuple à une ou deux dimentions")
		if type(data[0]) in (list, tuple):
			# Si data a deux dimentions, on boucle car la mathode append ne les gère pas
			for item in data:
				self.ws.append(item)
		else:
			self.ws.append(data)



	def auto_size(self, column_start=1, column_end=100, row_start=1, row_end=100):
		""" Redimentionnement des colonnes à la taille du contenu
			Par défaut pour les 100 premières colonnes et 100 premières lignes
		"""
		for column in range(column_start, column_end + 1):
			max_width = 10
			for row in range(row_start, row_end + 1):
				width = len(str(self.ws.cell(row, column).value)) * 1.1
				if width > max_width:
					max_width = width
			self.ws.column_dimensions[get_column_letter(column)].width = max_width


		# for column_cells in self.ws.columns:
		#     length = max(len(str(cell.value)) for cell in column_cells)
		#     self.ws.column_dimensions[column_cells[0].column_letter].width = length

	def color_row(self, row_number, color, fill_type = "solid"):
		""" Colorer une ligne """
		for rows in self.ws.iter_cols(min_row=row_number, max_row=row_number, min_col=None, max_col=None):
			for cell in rows:
				cell.fill = PatternFill(start_color=color, end_color=color, fill_type = fill_type)


	def close(self):
		self.wb.save(self.filepath)
		self.wb.close()
